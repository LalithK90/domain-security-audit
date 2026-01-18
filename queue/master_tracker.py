#!/usr/bin/env python3
"""Master Excel tracker - consolidates all domain scan results into a single Excel file."""

import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MASTER_FILE = Path(__file__).parent.parent / 'reports' / 'master_tracker.xlsx'
QUEUE_FILE = Path(__file__).parent / 'domain_queue.json'


def init_master_tracker():
    """Initialize master tracker Excel file with all required sheets."""
    MASTER_FILE.parent.mkdir(parents=True, exist_ok=True)
    
    if MASTER_FILE.exists():
        print(f"✓ Master tracker already exists: {MASTER_FILE}")
        return
    
    wb = Workbook()
    
    # Sheet 1: All Domains Summary
    ws_summary = wb.active
    ws_summary.title = "All Domains Summary"
    headers = ['Domain', 'Status', 'Total Score', 'Risk Rating', 'Subdomains Found', 
               'Active Subdomains', 'Scanned At', 'Report Path']
    ws_summary.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(1, col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.column_dimensions[get_column_letter(col_num)].width = 18
    
    # Sheet 2: Queue Status
    ws_queue = wb.create_sheet("Queue Status")
    queue_headers = ['Domain', 'Queue State', 'Added At']
    ws_queue.append(queue_headers)
    
    for col_num, header in enumerate(queue_headers, 1):
        cell = ws_queue.cell(1, col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_queue.column_dimensions[get_column_letter(col_num)].width = 25
    
    # Sheet 3: Scan History
    ws_history = wb.create_sheet("Scan History")
    history_headers = ['Timestamp', 'Domain', 'Action', 'Result', 'Details']
    ws_history.append(history_headers)
    
    for col_num, header in enumerate(history_headers, 1):
        cell = ws_history.cell(1, col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_history.column_dimensions[get_column_letter(col_num)].width = 20
    
    wb.save(MASTER_FILE)
    print(f"✓ Created master tracker: {MASTER_FILE}")


def add_scan_result(domain: str, total_score: float, risk_rating: str, 
                    subdomains_found: int, active_subdomains: int, 
                    report_path: str, status: str = "Completed"):
    """Add a domain scan result to the master tracker."""
    if not MASTER_FILE.exists():
        init_master_tracker()
    
    wb = load_workbook(MASTER_FILE)
    ws = wb["All Domains Summary"]
    
    # Check if domain already exists
    domain_row = None
    for row_num in range(2, ws.max_row + 1):
        if ws.cell(row_num, 1).value == domain:
            domain_row = row_num
            break
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    new_data = [domain, status, total_score, risk_rating, subdomains_found, 
                active_subdomains, timestamp, report_path]
    
    if domain_row:
        # Update existing row
        for col_num, value in enumerate(new_data, 1):
            ws.cell(domain_row, col_num).value = value
        action = "Updated"
    else:
        # Append new row
        ws.append(new_data)
        domain_row = ws.max_row
        action = "Added"
    
    # Color code risk rating
    risk_cell = ws.cell(domain_row, 4)
    if risk_rating == "CRITICAL":
        risk_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        risk_cell.font = Font(color="FFFFFF", bold=True)
    elif risk_rating == "HIGH":
        risk_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    elif risk_rating == "MEDIUM":
        risk_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    elif risk_rating == "LOW":
        risk_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    # Add to scan history
    ws_history = wb["Scan History"]
    ws_history.append([timestamp, domain, "Scan Complete", status, 
                      f"Score: {total_score}, Risk: {risk_rating}"])
    
    wb.save(MASTER_FILE)
    print(f"✓ {action} {domain} in master tracker: {total_score}% ({risk_rating})")


def update_queue_status():
    """Update queue status sheet with current queue state."""
    if not QUEUE_FILE.exists():
        return
    
    if not MASTER_FILE.exists():
        init_master_tracker()
    
    with open(QUEUE_FILE, 'r') as f:
        queue = json.load(f)
    
    wb = load_workbook(MASTER_FILE)
    ws = wb["Queue Status"]
    
    # Clear existing data (keep headers)
    ws.delete_rows(2, ws.max_row)
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Add pending domains
    for domain in queue.get('domains', []):
        ws.append([domain, "Pending", timestamp])
    
    # Add completed domains
    for domain in queue.get('completed', []):
        ws.append([domain, "Completed", timestamp])
        # Color completed rows green
        row = ws.max_row
        for col in range(1, 4):
            ws.cell(row, col).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    # Add failed domains
    for failed in queue.get('failed', []):
        domain = failed['domain'] if isinstance(failed, dict) else failed
        ws.append([domain, "Failed", timestamp])
        # Color failed rows red
        row = ws.max_row
        for col in range(1, 4):
            ws.cell(row, col).fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    
    wb.save(MASTER_FILE)
    print(f"✓ Updated queue status in master tracker")


def add_scan_history(domain: str, action: str, result: str, details: str = ""):
    """Add an entry to scan history."""
    if not MASTER_FILE.exists():
        init_master_tracker()
    
    wb = load_workbook(MASTER_FILE)
    ws = wb["Scan History"]
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, domain, action, result, details])
    
    wb.save(MASTER_FILE)


def get_summary() -> Dict[str, Any]:
    """Get summary statistics from master tracker."""
    if not MASTER_FILE.exists():
        return {
            'total_scanned': 0,
            'pending': 0,
            'completed': 0,
            'failed': 0,
            'avg_score': 0,
            'risk_breakdown': {}
        }
    
    wb = load_workbook(MASTER_FILE)
    ws = wb["All Domains Summary"]
    
    total = ws.max_row - 1  # Excluding header
    scores = []
    risk_counts = {}
    
    for row in range(2, ws.max_row + 1):
        score = ws.cell(row, 3).value
        risk = ws.cell(row, 4).value
        
        if score is not None:
            scores.append(float(score))
        
        if risk:
            risk_counts[risk] = risk_counts.get(risk, 0) + 1
    
    # Get queue stats
    if QUEUE_FILE.exists():
        with open(QUEUE_FILE, 'r') as f:
            queue = json.load(f)
            pending = len(queue.get('domains', []))
            completed = len(queue.get('completed', []))
            failed = len(queue.get('failed', []))
    else:
        pending = completed = failed = 0
    
    return {
        'total_scanned': total,
        'pending': pending,
        'completed': completed,
        'failed': failed,
        'avg_score': sum(scores) / len(scores) if scores else 0,
        'risk_breakdown': risk_counts
    }


if __name__ == '__main__':
    if len(sys.argv) > 1:
        command = sys.argv[1]
        
        if command == 'init':
            init_master_tracker()
        
        elif command == 'update-queue':
            update_queue_status()
        
        elif command == 'summary':
            stats = get_summary()
            print("\n" + "=" * 60)
            print("📊 MASTER TRACKER SUMMARY")
            print("=" * 60)
            print(f"Total Scanned:  {stats['total_scanned']}")
            print(f"Queue Pending:  {stats['pending']}")
            print(f"Completed:      {stats['completed']}")
            print(f"Failed:         {stats['failed']}")
            print(f"Average Score:  {stats['avg_score']:.1f}%")
            
            if stats['risk_breakdown']:
                print("\nRisk Breakdown:")
                for risk, count in sorted(stats['risk_breakdown'].items()):
                    print(f"  {risk}: {count}")
            print("=" * 60)
        
        elif command == 'add':
            # Test: python master_tracker.py add domain.com 85.5 HIGH 50 30 /path/to/report.xlsx
            if len(sys.argv) >= 8:
                add_scan_result(sys.argv[2], float(sys.argv[3]), sys.argv[4],
                               int(sys.argv[5]), int(sys.argv[6]), sys.argv[7])
    else:
        # Default: show summary
        init_master_tracker()
        update_queue_status()
        stats = get_summary()
        print(f"\n✓ Master tracker ready: {MASTER_FILE}")
        print(f"  Scanned: {stats['total_scanned']} | Pending: {stats['pending']} | Avg Score: {stats['avg_score']:.1f}%")
