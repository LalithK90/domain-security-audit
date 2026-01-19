"""XLSX Seed Loader - Load existing security reports as initial seeds.

WHY THIS EXISTS:
- Previous manual scan reports contain validated subdomains
- Real source attribution ("xlsx_seed") for reproducibility
- Jump-start enumeration with known-good targets
- Avoids duplicate work

EXPECTED FORMATS:
1. *_security_report.xlsx with "Subdomains" or "Targets" sheet
2. master_tracker.xlsx with domain inventory
3. Any Excel with FQDN column
"""

import logging
from pathlib import Path
from typing import List, Set, Optional
import re

logger = logging.getLogger(__name__)


def load_xlsx_seeds(domain: str, search_paths: Optional[List[Path]] = None) -> Set[str]:
    """Load subdomains from existing XLSX security reports.
    
    Args:
        domain: Base domain to filter results
        search_paths: List of paths to search for XLSX files. Defaults to:
                     - ./reports/
                     - ./out/{domain}/
                     - ./src/legacy/reports/
    
    Returns:
        Set of discovered FQDNs with source attribution
        
    WHY: Real source attribution for research reproducibility.
    """
    # Try importing openpyxl
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed - XLSX seed loading disabled")
        logger.warning("Install with: pip install openpyxl")
        return set()
    
    # Default search paths
    if not search_paths:
        search_paths = [
            Path("reports"),
            Path("out") / domain,
            Path("src/legacy/reports"),
            Path("."),  # Current directory
        ]
    
    discovered = set()
    files_found = 0
    
    # Search for XLSX files
    for search_path in search_paths:
        if not search_path.exists():
            continue
            
        for xlsx_file in search_path.rglob("*.xlsx"):
            # Skip temp files
            if xlsx_file.name.startswith("~$") or xlsx_file.name.startswith("."):
                continue
                
            try:
                logger.info(f"Loading seeds from {xlsx_file.name}...")
                subdomains = _extract_domains_from_xlsx(xlsx_file, domain)
                
                if subdomains:
                    discovered.update(subdomains)
                    files_found += 1
                    logger.info(f"  Found {len(subdomains)} subdomains in {xlsx_file.name}")
                    
            except Exception as e:
                logger.warning(f"Failed to load {xlsx_file.name}: {e}")
                continue
    
    if files_found > 0:
        logger.info(f"XLSX Seed Loading: {len(discovered)} unique subdomains from {files_found} files")
    else:
        logger.info("XLSX Seed Loading: No Excel files found in search paths")
    
    return discovered


def _extract_domains_from_xlsx(xlsx_path: Path, domain: str) -> Set[str]:
    """Extract FQDNs from an Excel file.
    
    Strategy:
    1. Look for sheets named: Subdomains, Targets, Results, Scan Results
    2. Look for columns containing: fqdn, subdomain, domain, host, target
    3. Extract and validate all values that end with base domain
    
    Args:
        xlsx_path: Path to Excel file
        domain: Base domain to filter
        
    Returns:
        Set of valid FQDNs
    """
    from openpyxl import load_workbook
    
    discovered = set()
    
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        
        # Target sheet names (priority order)
        target_sheets = ['Subdomains', 'Targets', 'Results', 'Scan Results', 'Data', 'Summary']
        
        # Try target sheets first
        sheets_to_check = []
        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                sheets_to_check.append(wb[sheet_name])
        
        # If no target sheets, check all sheets
        if not sheets_to_check:
            sheets_to_check = [wb[sheet] for sheet in wb.sheetnames]
        
        for sheet in sheets_to_check:
            # Find header row (usually row 1)
            headers = []
            for cell in sheet[1]:
                val = str(cell.value).lower() if cell.value else ""
                headers.append(val)
            
            # Find columns that might contain FQDNs
            target_columns = []
            for idx, header in enumerate(headers):
                if any(keyword in header for keyword in ['fqdn', 'subdomain', 'domain', 'host', 'target', 'name']):
                    target_columns.append(idx)
            
            # If no obvious columns, check all columns
            if not target_columns:
                target_columns = list(range(len(headers)))
            
            # Extract FQDNs from target columns
            for row in sheet.iter_rows(min_row=2, values_only=True):
                for col_idx in target_columns:
                    if col_idx >= len(row):
                        continue
                        
                    value = row[col_idx]
                    if not value:
                        continue
                    
                    # Clean and validate
                    fqdn = str(value).strip().lower()
                    
                    # Basic FQDN validation
                    if _is_valid_fqdn(fqdn, domain):
                        discovered.add(fqdn)
        
        wb.close()
        
    except Exception as e:
        logger.debug(f"Error parsing {xlsx_path.name}: {e}")
        raise
    
    return discovered


def _is_valid_fqdn(fqdn: str, domain: str) -> bool:
    """Validate that a string is a valid FQDN for the given domain.
    
    Args:
        fqdn: Candidate FQDN
        domain: Base domain
        
    Returns:
        True if valid FQDN ending with domain
    """
    # Remove protocol if present
    fqdn = re.sub(r'^https?://', '', fqdn)
    
    # Remove path if present
    fqdn = fqdn.split('/')[0]
    
    # Remove port if present
    fqdn = fqdn.split(':')[0]
    
    # Remove trailing dot
    fqdn = fqdn.rstrip('.')
    
    # Must end with domain
    if not fqdn.endswith(domain):
        return False
    
    # Must be valid DNS name format
    if not re.match(r'^[a-z0-9]([a-z0-9-]{0,61}[a-z0-9])?(\.[a-z0-9]([a-z0-9-]{0,61}[a-z0-9])?)*$', fqdn):
        return False
    
    # Must contain only valid characters
    if not all(c.isalnum() or c in '.-' for c in fqdn):
        return False
    
    return True
